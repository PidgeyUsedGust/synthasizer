import pytest
import openpyxl
from pathlib import Path
from synthasizer.table import detect

data = Path(__file__).parent / "data"


@pytest.fixture(scope="module")
def nba():
    return detect(openpyxl.load_workbook(data / "nba.xlsx")["Sheet1"])[0]


@pytest.fixture(scope="module")
def nurse():
    return detect(openpyxl.load_workbook(data / "nurse.xlsx")["nurse"])[0]


@pytest.fixture(scope="module")
def nurse2():
    return detect(openpyxl.load_workbook(data / "nurse_weird.xlsx")["Sheet1"])[0]


@pytest.fixture(scope="module")
def reden():
    return detect(openpyxl.load_workbook(data / "reden.xlsx")["Sheet1"])[0]


@pytest.fixture(scope="module")
def deeltijdswerk():
    return detect(openpyxl.load_workbook(data / "deeltijdswerk.xlsx")["Sheet1"])[0]


@pytest.fixture(scope="module")
def part():
    return detect(openpyxl.load_workbook(data / "parttime.xlsx")["default"])[0]


@pytest.fixture(scope="module")
def part_e():
    return detect(openpyxl.load_workbook(data / "parttime.xlsx")["extended"])[0]


@pytest.fixture(scope="module")
def icecream():
    return detect(openpyxl.load_workbook(data / "icecream.xlsx")["month"])[0]


@pytest.fixture(scope="module")
def icecreamyear():
    return detect(openpyxl.load_workbook(data / "icecream.xlsx")["year"])[0]


@pytest.fixture(scope="module")
def car():
    return detect(openpyxl.load_workbook(data / "car.xlsx")["Sheet1"])[0]
