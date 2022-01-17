"""
Table wrapper with support for colorings as they
need to be propagated through the table after
applying transformations.
"""
import itertools
import pandas as pd
import numpy as np
from copy import copy
from collections import defaultdict
from typing import List, Any, Optional, Tuple
from functools import cached_property
from operator import attrgetter, itemgetter
from openpyxl.cell.cell import Cell as PyxlCell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, range_boundaries
from synthasizer.utilities import infer_types, nothing


class Cell:
    """A spreadsheet cell.

    Style is represented as a simple dictionary
    that can be used to set any cell properties.

    Color is an integer defining the type of cell.
    It can also be used to define other criteria,
    such as a segmentation.

    """

    def __init__(self, value: Optional[Any] = None, **kwargs):
        self.value = none(value)
        self.color = 0
        self.base = None
        self.style = defaultdict(nothing)
        self.style.update(kwargs)

    def same_style(self, other: "Cell") -> bool:
        """Check if same style.

        Only returns true if exactly the same. Probably
        not very robust in practice.

        """
        if set(self.style) != set(other.style):
            return False
        return all(self.style[k] == v for k, v in other.style.items())

    @cached_property
    def datatype(self) -> str:
        return pd.api.types.infer_dtype([self.value])

    def with_color(self, color: int) -> "Cell":
        """Return cell with new color applied."""
        cell = copy(self)
        cell.color = color
        return cell

    def __eq__(self, other):
        return isinstance(other, Cell) and (self.value == other.value)

    def __ne__(self, other):
        return self.value != other.value

    def __lt__(self, other):
        return self.value < other.value

    def __gt__(self, other):
        return self.value > other.value

    def __bool__(self):
        return not pd.isna(self.value)

    def __copy__(self):
        cell = Cell(self.value)
        cell.style = copy(self.style)
        cell.color = self.color
        return cell

    def __hash__(self):
        return hash((self.value,))

    def __str__(self):
        if self.value is None:
            return ""
        return str(self.value)

    def __repr__(self):
        return "Cell({})".format(repr(self.value))

    @classmethod
    def from_openpyxl(cls, cell: PyxlCell) -> "Cell":
        # extract styles from cell
        style = dict(
            bold=cell.font.bold,
            italic=cell.font.italic,
            underline=cell.font.underline,
            size=cell.font.size,
            fill=cell.fill.fgColor.value,
        )
        # sometimes font color is not set
        if cell.font.color is not None:
            style["color"] = cell.font.color.value
        # build cell
        new_cell = cls(cell.value, **style)
        return new_cell


class Table:
    """Table wrapper with colors.

    Colors are cell properties that are propagated
    while wrangling, so we need a wrapper.

    """

    def __init__(self, df: Optional[pd.DataFrame] = None):
        """Initialise table.

        Args:
            df: A pandas dataframe.

        """
        self.df = pd.DataFrame(dtype=object)
        self.current_color = 1
        self.header = 0
        if df is not None:
            data = [[Cell(v) for v in row] for _, row in df.iterrows()]
            columns = [Cell(v) for v in df.columns]
            self.df = pd.DataFrame(data, columns=columns, dtype=object)

    def color(self, x: int, y: int, color: int = 0) -> "Table":
        """Color one cell.

        Args:
            x, y: Column and row of the cell to be colored.
            color: Color to use. If zero, create a new color.

        Returns:
            A new table with this cell marked in a new color.

        """
        table = copy(self)
        if table.df.iat[y, x].color == 0:
            if color == 0:
                color = table.current_color
                table.current_color += 1
            table.df.iat[y, x] = table.df.iat[y, x].with_color(color)
        return table

    def color_all(self, positions: List[Tuple[int, int]], colors: List[int] = None):
        """Color multiple cells.

        Args:
            positions: List of (x, y) positions of cells to
                be colored.
            colors: List of colors to assign to the cells.

        Returns:
            A new Table with all given cells marked in a
            different color.

        """
        if colors is None:
            colors = [0] * len(positions)
        table = self
        for i, position in enumerate(positions):
            table = table.color(*position, color=colors[i])
        return table

    @cached_property
    def color_df(self) -> pd.DataFrame:
        """Colors as dataframe.

        Headers are encoded as a single index with
        tuples of colors, as this makes working with
        color dataframes significantly easier.

        """
        df = pd.DataFrame(np.vectorize(attrgetter("color"))(self.df.values))
        if isinstance(self.df.columns, pd.MultiIndex):
            df.columns = [tuple(cell.color for cell in c) for c in self.df.columns]
        else:
            df.columns = [(c.color,) for c in self.df.columns]
        return df

    @cached_property
    def n_colors(self) -> int:
        """Number of colors."""
        color_df = self.color_df
        colors = set(color_df.values.ravel("k"))
        if self.header > 0:
            colors_header = set(itertools.chain.from_iterable(color_df.columns))
        else:
            colors_header = set()
        return len((colors | colors_header) - {0})

    @cached_property
    def cell_types(self) -> np.ndarray:
        """Get the types of all cells as a numpy matrix.

        Returns:
            A two-dimensional matrix containing the typs
            of each cell.

        """
        return np.vectorize(attrgetter("datatype"))(self.df.values)

    @cached_property
    def column_types(self) -> List[str]:
        """Infer the type of each column.

        Returns:
            A list of length `self.width` with the
            inferred type of each column. See
            `utilities.infer_types` for more information.

        """
        return [infer_types(column) for column in self.cell_types.T]

    @cached_property
    def dataframe(self) -> pd.DataFrame:
        """Unwrapped dataframe.

        Warning, this method is slow. Use
        it sparingly.

        """
        return self.df.applymap(lambda cell: cell.value).convert_dtypes()

    @property
    def cells(self) -> List[Cell]:
        """Get all cells in the table."""
        cells = self.df.values.ravel("K")
        cells = cells[cells != Cell(None)]
        if self.header > 0:
            if isinstance(self.df.columns, pd.MultiIndex):
                cells_header = itertools.chain.from_iterable(self.df.columns)
            else:
                cells_header = self.df.columns
            cells_header = list(filter(bool, cells_header))
            cells = np.hstack((cells, cells_header))
        return list(cells)

    @property
    def height(self) -> int:
        return len(self.df.index)

    @property
    def width(self) -> int:
        return len(self.df.columns)

    @classmethod
    def from_csv(cls, file: str, header: Optional[int] = None) -> "Table":
        """Load from CSV."""
        return Table(pd.read_csv(file, header=header))

    @classmethod
    def from_openpyxl(cls, data: List[List[PyxlCell]]) -> "Table":
        """Load from openpyxl cells."""
        cells = [[Cell.from_openpyxl(c) for c in row] for row in data]
        index = [Cell(i) for i in range(len(cells[0]))]
        table = Table()
        table.df = pd.DataFrame(cells, columns=index)
        return table

    def __getitem__(self, i):
        """Implement slicing.

        By default, forward everything to iloc, except
        for single integers, which are considered
        to be columns.

        """
        if isinstance(i, int):
            return self.df.iloc[:, i]
        return self.df.iloc[i]

    def __copy__(self):
        new = Table()
        new.df = self.df.copy()
        new.current_color = self.current_color
        new.header = self.header
        return new

    def __str__(self):
        return str(self.df)

    def __repr__(self):
        return str(self.df)

    def __hash__(self) -> int:
        hash_header = pd.util.hash_pandas_object(self.df.columns)
        return hash((self.df.values.tobytes(), hash_header.values.tobytes()))


na_values = {
    "",
    "-1.#IND",
    "1.#QNAN",
    "1.#IND",
    "-1.#QNAN",
    "#N/A",
    "N/A",
    "NA",
    "#NA",
    "NULL",
    "NaN",
    "-NaN",
    "nan",
    "-nan",
    "<NA>",
}
"""Default NA values from pandas."""


def none(value: Any):
    """Check if value is nan."""
    if (value in na_values) or pd.isna(value):
        return None
    return value


def detect(sheet: Worksheet, min_size: int = 4) -> List[Table]:
    """Detect tables in a worksheet.

    Args:
        min_size: Number of cells that a table should contain.

    Returns:
        A list of tables.

    """
    unmerge(sheet)
    tables = list()
    for i in detect_ranges(sheet):
        min_col, min_row, max_col, max_row = range_boundaries(i)
        if (max_row - min_row) * (max_col - min_col) >= min_size:
            tables.append(Table.from_openpyxl(sheet[i]))
    return tables


def detect_ranges(sheet: Worksheet) -> List[str]:
    """Extract tables from a worksheet."""
    # initialise mask
    mask = np.zeros((sheet.max_row, sheet.max_column), dtype=int)
    for row in sheet.rows:
        for cell in row:
            if cell.value is not None:
                # print(cell.row, cell.column, cell.value)
                mask[cell.row - 1, cell.column - 1] = 1.0
    # look for tables
    tables = list()
    while np.count_nonzero(mask) > 0:
        # find nonzero cell
        point = next(zip(*np.where(mask > 0)))
        # detect connected region
        x1, y1, x2, y2 = detect_from(mask, point)
        # add to results
        tables.append("{}:{}".format(to_excel((y1, x1)), to_excel((y2, x2))))
        # zero out the discovered region
        mask[x1 : x2 + 1, y1 : y2 + 1] = 0
    return tables


def detect_from(table: np.ndarray, index: Tuple[int, int]) -> str:
    """Detect table from a given index."""
    queue = [index]
    cells = set()
    while len(queue) > 0:
        a, b = queue.pop()
        for dx, dy in itertools.product((-1, 0, 1), repeat=2):
            if a + dx >= 0 and b + dy >= 0:
                try:
                    if table[a + dx, b + dy] != 0:
                        if (a + dx, b + dy) not in cells:
                            queue.append((a + dx, b + dy))
                            cells.add(((a + dx, b + dy)))
                except IndexError:
                    pass
    return (
        min(cells, key=itemgetter(0))[0],
        min(cells, key=itemgetter(1))[1],
        max(cells, key=itemgetter(0))[0],
        max(cells, key=itemgetter(1))[1],
    )


def to_excel(point: Tuple[int, int]) -> str:
    """Make into Excel coordinate.

    Args:
        point: A one-indexed (column, row) tuple.

    """
    return get_column_letter(point[0] + 1) + str(point[1] + 1)


def unmerge(sheet: Worksheet) -> None:
    """Unmerge all cells in a worksheet."""
    for i in list(sheet.merged_cells.ranges):
        i = str(i)
        value = next((c for r in sheet[i] for c in r if c.value is not None), None)
        sheet.unmerge_cells(i)
        if value is not None:
            for row in sheet[i]:
                for cell in row:
                    cell.value = value.value
                    if value.has_style:
                        cell.font = copy(value.font)
                        cell.border = copy(value.border)
                        cell.fill = copy(value.fill)
                        cell.number_format = copy(value.number_format)
                        cell.protection = copy(value.protection)
                        cell.alignment = copy(value.alignment)


def trim(sheet: Worksheet, table: str) -> str:
    """Trim exterior rows and columns that contain only a single element."""

    min_col, min_row, max_col, max_row = range_boundaries(table)
